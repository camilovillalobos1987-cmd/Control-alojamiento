from flask import (
    Flask, render_template, request, redirect, url_for,
    jsonify, send_file, flash, abort, session
)
import io
import csv
import json
from datetime import date, datetime, timezone, timedelta

# Zona horaria Chile (UTC-4 invierno / UTC-3 verano)
CHILE_TZ = timezone(timedelta(hours=-4))

MESES_ES = [
    'enero','febrero','marzo','abril','mayo','junio',
    'julio','agosto','septiembre','octubre','noviembre','diciembre'
]


def _fecha_larga(valor) -> str:
    """Convierte 'YYYY-MM-DD' → '17 de abril de 2026'"""
    if not valor:
        return '—'
    try:
        d = date.fromisoformat(str(valor))
        return f"{d.day} de {MESES_ES[d.month - 1]} de {d.year}"
    except Exception:
        return str(valor)

from config import SECRET_KEY, NOMBRE_EMPRESA, TURNOS, CONTRATURNOS, GRUPOS_TURNOS
import database as db
from qr_manager import generar_token_qr, generar_imagen_qr, generar_qr_base64
from turnos import calcular_estado_turno, trabajadores_que_llegan_pronto
from email_service import notificar_llegada, notificar_cambio_habitacion
from auth import (
    login_required, admin_required,
    login_user, logout_user, check_password,
)

app = Flask(__name__)
app.secret_key = SECRET_KEY

# ── INICIALIZACIÓN AUTOMÁTICA DE BASE DE DATOS (Para Producción / Gunicorn) ──
from init_data import init as db_init
from database import migrar_db
db_init()
migrar_db()


# ═══════════════════════════════════════════════════════
# CONTEXT PROCESSOR
# ═══════════════════════════════════════════════════════
@app.template_filter('fecha_cl')
def fecha_cl_filter(valor):
    return _fecha_larga(valor)


@app.context_processor
def inject_globals():
    ahora_chile = datetime.now(CHILE_TZ)
    return {
        "NOMBRE_EMPRESA": NOMBRE_EMPRESA,
        "hoy": ahora_chile.strftime("%d/%m/%Y"),
        "hoy_iso": ahora_chile.strftime("%Y-%m-%d"),
        "hora_chile": ahora_chile.strftime("%H:%M"),
    }


# ═══════════════════════════════════════════════════════
# AUTENTICACIÓN
# ═══════════════════════════════════════════════════════
@app.route("/login", methods=["GET", "POST"])
def login():
    from auth import current_user
    if current_user():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = db.get_usuario_by_username(username)
        if user and check_password(password, user["password"]):
            login_user(user)
            flash(f"Bienvenido, {user.get('nombre') or username}.", "success")
            return redirect(url_for("dashboard"))
        flash("Usuario o contraseña incorrectos. Verifica tus credenciales.", "danger")
    return render_template("auth/login.html")


@app.route("/logout")
def logout():
    logout_user()
    flash("Sesión cerrada correctamente. Hasta pronto.", "success")
    return redirect(url_for("login"))


@app.errorhandler(403)
def forbidden(_e):
    return render_template("auth/403.html"), 403


# ═══════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════
def _agregar_metricas_ajustadas(metricas, trabajadores):
    from turnos import calcular_estado_turno
    estados_excepcion = {'Licencia Médica','Vacaciones','Permiso','Falla'}
    hoy_str = date.today().isoformat()
    
    # Recalculate true states in memory
    for t in trabajadores:
        if t["estado"] in estados_excepcion:
            novedad = db.get_novedad_vigente(t["id"])
            if novedad and novedad.get("fecha_inicio") and novedad["fecha_inicio"] > hoy_str:
                if t.get("turno") and t.get("fecha_inicio_ciclo"):
                    info = calcular_estado_turno(t["turno"], t["fecha_inicio_ciclo"])
                    t["estado"] = info["estado_calculado"]

    # Rebuild por_turno distribution logically
    turnos_dict = {}
    for t in trabajadores:
        if t["estado"] == "Desvinculado" or not t.get("turno"):
            continue
        turno = t["turno"]
        if turno not in turnos_dict:
            turnos_dict[turno] = {"turno": turno, "total": 0, "en_faena": 0}
        turnos_dict[turno]["total"] += 1
        if t["estado"] == "Activo en campamento":
            turnos_dict[turno]["en_faena"] += 1
            
    metricas["por_turno"] = list(turnos_dict.values())
    return metricas

@app.route("/")
@login_required
def dashboard():
    metricas = db.get_metricas_dashboard()
    trabajadores = db.get_todos_trabajadores()
    
    metricas = _agregar_metricas_ajustadas(metricas, trabajadores)
    
    llegadas = trabajadores_que_llegan_pronto(trabajadores, horas=48)
    habitaciones = db.get_todas_habitaciones()
    return render_template(
        "index.html",
        metricas=metricas,
        llegadas=llegadas,
        habitaciones=habitaciones,
    )


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    metricas = db.get_metricas_dashboard()
    trabajadores = db.get_todos_trabajadores()
    metricas = _agregar_metricas_ajustadas(metricas, trabajadores)
    return jsonify(metricas)


# ═══════════════════════════════════════════════════════
# TRABAJADORES
# ═══════════════════════════════════════════════════════
@app.route("/trabajadores")
@admin_required
def trabajadores_lista():
    estado_filtro = request.args.get("estado", "")
    turno_filtro = request.args.get("turno", "")
    busqueda = request.args.get("q", "").lower()

    todos = db.get_todos_trabajadores()

    # Ajustar estado real para novedades futuras
    estados_excepcion = {'Licencia Médica','Vacaciones','Permiso','Falla'}
    from datetime import date as _date
    hoy_str = _date.today().isoformat()

    for t in todos:
        if t["estado"] in estados_excepcion:
            novedad = db.get_novedad_vigente(t["id"])
            if novedad and novedad.get("fecha_inicio") and novedad["fecha_inicio"] > hoy_str:
                # Novedad futura: calculamos ciclo normal para mostrar
                if t.get("turno") and t.get("fecha_inicio_ciclo"):
                    info = calcular_estado_turno(t["turno"], t["fecha_inicio_ciclo"])
                    t["estado"] = info["estado_calculado"]

    filtrados = [
        t for t in todos
        if (not estado_filtro or t["estado"] == estado_filtro)
        and (not turno_filtro or t["turno"] == turno_filtro)
        and (not busqueda or busqueda in t["nombre"].lower() or busqueda in t["rut"].lower())
    ]

    PER_PAGE = 20
    total = len(filtrados)
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page = max(1, min(int(request.args.get("page", 1)), total_pages))
    start = (page - 1) * PER_PAGE
    paginados = filtrados[start:start + PER_PAGE]

    return render_template(
        "trabajadores/lista.html",
        trabajadores=paginados,
        total=total,
        page=page,
        total_pages=total_pages,
        estado_filtro=estado_filtro,
        turno_filtro=turno_filtro,
        busqueda=busqueda,
        turnos=list(TURNOS.keys()),
        estados=["Activo en campamento", "En descanso", "Permiso",
                 "Falla", "Licencia Médica", "Vacaciones", "Desvinculado"],
    )


@app.route("/trabajadores/nuevo", methods=["GET", "POST"])
@admin_required
def trabajador_nuevo():
    if request.method == "POST":
        data = {
            "nombre": request.form["nombre"].strip(),
            "rut": request.form["rut"].strip(),
            "cargo": request.form.get("cargo", "").strip(),
            "turno": request.form["turno"],
            "email": request.form.get("email", "").strip(),
            "estado": request.form.get("estado", "En descanso"),
            "fecha_inicio_ciclo": request.form.get("fecha_inicio_ciclo") or None,
        }
        try:
            id_nuevo = db.crear_trabajador(data)
            _regenerar_qr(id_nuevo)
            flash("Trabajador creado exitosamente.", "success")
            return redirect(url_for("trabajador_detalle", id=id_nuevo))
        except Exception as e:
            flash(f"Error al crear trabajador: {e}", "danger")

    return render_template("trabajadores/formulario.html",
                           trabajador=None, turnos=list(TURNOS.keys()),
                           turnos_info={k: {"ref_inicio": v.get("ref_inicio"), "contraturno": CONTRATURNOS.get(k)} for k, v in TURNOS.items()},
                           estados=["Activo en campamento", "En descanso", "Permiso",
                                    "Falla", "Licencia Médica", "Vacaciones", "Desvinculado"])


@app.route("/trabajadores/<int:id>")
@admin_required
def trabajador_detalle(id):
    t = db.get_trabajador(id)
    if not t:
        abort(404)

    turno_info = {}
    estados_excepcion = {'Licencia Médica','Vacaciones','Permiso','Falla','Desvinculado'}
    
    # Obtener novedad si el estado es de excepción
    novedad = None
    es_novedad_activa = False
    es_novedad_futura = False
    
    if t["estado"] in estados_excepcion:
        novedad = db.get_novedad_vigente(t["id"])
        if novedad and novedad.get("fecha_inicio"):
            from datetime import date as _date
            hoy_str = _date.today().isoformat()
            if novedad["fecha_inicio"] > hoy_str:
                es_novedad_futura = True
                # Override the state temporarily for the header badge
                if t.get("turno") and t.get("fecha_inicio_ciclo"):
                    info = calcular_estado_turno(t["turno"], t["fecha_inicio_ciclo"])
                    t["estado"] = info["estado_calculado"]
            else:
                es_novedad_activa = True
                
            if novedad.get("fecha_fin"):
                try:
                    d0 = _date.fromisoformat(novedad["fecha_inicio"])
                    d1 = _date.fromisoformat(novedad["fecha_fin"])
                    novedad["duracion_dias"] = (d1 - d0).days + 1
                except Exception:
                    novedad["duracion_dias"] = None
        else:
            # Si no hay fechas o novedad, asumimos activa por defecto
            es_novedad_activa = True
            
    # Si la novedad es a futuro, o si no hay novedad de excepción, calculamos el ciclo normal para HOY
    if (t.get("turno") and t.get("fecha_inicio_ciclo")) and (t["estado"] not in estados_excepcion or es_novedad_futura):
        turno_info = calcular_estado_turno(t["turno"], t["fecha_inicio_ciclo"])

    movimientos = db.get_movimientos(limit=20, trabajador_id=id)
    habitaciones_disp = db.get_habitaciones_disponibles()
    qr_b64 = None
    if t.get("qr_token") and not t.get("qr_revocado"):
        qr_b64 = generar_qr_base64(t["qr_token"])
    return render_template(
        "trabajadores/detalle.html",
        t=t,
        turno_info=turno_info,
        novedad=novedad,
        es_novedad_activa=es_novedad_activa,
        es_novedad_futura=es_novedad_futura,
        movimientos=movimientos,
        habitaciones_disponibles=habitaciones_disp,
        qr_b64=qr_b64,
    )


@app.route("/trabajadores/<int:id>/editar", methods=["GET", "POST"])
@admin_required
def trabajador_editar(id):
    t = db.get_trabajador(id)
    if not t:
        abort(404)
    if request.method == "POST":
        data = {
            "nombre": request.form["nombre"].strip(),
            "rut": request.form["rut"].strip(),
            "cargo": request.form.get("cargo", "").strip(),
            "turno": request.form["turno"],
            "email": request.form.get("email", "").strip(),
            "estado": request.form.get("estado", t["estado"]),
            "fecha_inicio_ciclo": request.form.get("fecha_inicio_ciclo") or None,
        }
        db.actualizar_trabajador(id, data)
        _regenerar_qr(id)
        flash("Trabajador actualizado.", "success")
        return redirect(url_for("trabajador_detalle", id=id))
    return render_template("trabajadores/formulario.html",
                           trabajador=t, turnos=list(TURNOS.keys()),
                           turnos_info={k: {"ref_inicio": v.get("ref_inicio"), "contraturno": CONTRATURNOS.get(k)} for k, v in TURNOS.items()},
                           estados=["Activo en campamento", "En descanso", "Permiso",
                                    "Falla", "Licencia Médica", "Vacaciones", "Desvinculado"])


@app.route("/trabajadores/<int:id>/estado", methods=["POST"])
@admin_required
def trabajador_cambio_estado(id):
    t = db.get_trabajador(id)
    if not t:
        abort(404)

    nuevo_estado = request.form["estado"]
    liberar = request.form.get("liberar_pieza") == "1"
    observacion = request.form.get("observacion", "")
    fecha_inicio = request.form.get("fecha_inicio") or date.today().isoformat()
    fecha_fin = request.form.get("fecha_fin") or None

    # Actualizar estado
    db.actualizar_estado_trabajador(id, nuevo_estado)

    # Desvinculado: siempre liberar pieza y revocar QR
    if nuevo_estado == "Desvinculado":
        db.liberar_habitacion_de_trabajador(id)
        db.revocar_qr(id)
        liberar = True

    # Liberar pieza si el admin lo indica
    elif liberar:
        db.liberar_habitacion_de_trabajador(id)
        _regenerar_qr(id)

    # Registrar novedad
    db.registrar_novedad({
        "trabajador_id": id,
        "tipo": nuevo_estado if nuevo_estado in [
            "Permiso", "Falla", "Licencia Médica", "Vacaciones"
        ] else "Otro",
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "pieza_liberada": 1 if liberar else 0,
        "observacion": observacion,
        "registrado_por": "Administrador",
    })

    flash(f"Estado actualizado a '{nuevo_estado}'.", "success")
    return redirect(url_for("trabajador_detalle", id=id))


@app.route("/trabajadores/<int:id>/qr/descargar")
@admin_required
def descargar_qr(id):
    t = db.get_trabajador(id)
    if not t or not t.get("qr_token") or t.get("qr_revocado"):
        abort(404)
    img_bytes = generar_imagen_qr(t["qr_token"])
    return send_file(
        io.BytesIO(img_bytes),
        mimetype="image/png",
        as_attachment=True,
        download_name=f"QR_{t['rut'].replace('.','').replace('-','')}.png",
    )


@app.route("/trabajadores/<int:id>/tarjeta")
@admin_required
def tarjeta_qr(id):
    t = db.get_trabajador(id)
    if not t:
        abort(404)
    qr_b64 = None
    if t.get("qr_token") and not t.get("qr_revocado"):
        qr_b64 = generar_qr_base64(t["qr_token"])
    return render_template("qr/tarjeta.html", t=t, qr_b64=qr_b64)


@app.route("/trabajadores/<int:id>/regenerar_qr", methods=["POST"])
@admin_required
def regenerar_qr_trabajador(id):
    _regenerar_qr(id)
    flash("Código QR regenerado exitosamente.", "success")
    return redirect(url_for("trabajador_detalle", id=id))


@app.route("/trabajadores/<int:id>/notificar", methods=["POST"])
@admin_required
def notificar_trabajador(id):
    t = db.get_trabajador(id)
    if not t:
        abort(404)
    ok = notificar_llegada(t)
    if ok:
        flash("Correo de notificación enviado exitosamente.", "success")
    else:
        flash("No se pudo enviar el correo. Por favor, revisa la configuración.", "danger")
    return redirect(url_for("trabajador_detalle", id=id))


# ═══════════════════════════════════════════════════════
# IMPORTACIÓN MASIVA + ASIGNACIÓN ALEATORIA
# ═══════════════════════════════════════════════════════
@app.route("/trabajadores/importar", methods=["GET", "POST"])
@admin_required
def trabajadores_importar():
    resultado = None
    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename.lower().endswith(".csv"):
            flash("Por favor sube un archivo .csv válido.", "danger")
            return redirect(url_for("trabajadores_importar"))

        content = archivo.read().decode("utf-8-sig")  # utf-8-sig maneja BOM de Excel
        reader = csv.DictReader(io.StringIO(content))
        creados, omitidos, errores = 0, 0, []

        for i, row in enumerate(reader, start=2):
            nombre = (row.get("nombre") or "").strip()
            rut    = (row.get("rut")    or "").strip()
            if not nombre or not rut:
                errores.append(f"Fila {i}: nombre y RUT son obligatorios.")
                continue

            turno = (row.get("turno") or "").strip() or None
            if turno and turno not in TURNOS:
                errores.append(f"Fila {i} ({nombre}): turno '{turno}' no válido. Opciones: {', '.join(TURNOS.keys())}")
                continue

            estados_validos = ["Activo en campamento", "En descanso", "Permiso",
                               "Falla", "Licencia Médica", "Vacaciones", "Desvinculado"]
            estado = (row.get("estado") or "En descanso").strip()
            if estado not in estados_validos:
                estado = "En descanso"

            data = {
                "nombre":             nombre,
                "rut":                rut,
                "cargo":              (row.get("cargo")  or "").strip(),
                "turno":              turno,
                "email":              (row.get("email")  or "").strip(),
                "estado":             estado,
                "fecha_inicio_ciclo": (row.get("fecha_inicio_ciclo") or "").strip() or None,
            }
            try:
                id_nuevo = db.crear_trabajador(data)
                _regenerar_qr(id_nuevo)
                creados += 1
            except Exception as e:
                if "UNIQUE" in str(e):
                    omitidos += 1
                else:
                    errores.append(f"Fila {i} ({nombre}): {e}")

        resultado = {"creados": creados, "omitidos": omitidos, "errores": errores}

    sin_hab = db.get_trabajadores_sin_habitacion()
    return render_template("trabajadores/importar.html",
                           resultado=resultado,
                           turnos=list(TURNOS.keys()),
                           sin_habitacion=len(sin_hab))


@app.route("/trabajadores/importar/plantilla")
@admin_required
def plantilla_csv():
    """Descarga plantilla CSV lista para completar en Excel."""
    filas = [
        ["nombre", "rut", "cargo", "turno", "email", "estado", "fecha_inicio_ciclo"],
        ["Juan Pérez González", "12.345.678-9", "Operador", "14x14",
         "juan@empresa.cl", "En descanso", "2026-04-01"],
        ["María González S.", "9.876.543-2", "Supervisora", "7x7",
         "maria@empresa.cl", "Activo en campamento", "2026-04-05"],
        ["Pedro Rojas M.", "11.222.333-4", "Mecánico", "5x2",
         "pedro@empresa.cl", "En descanso", "2026-04-03"],
    ]
    out = io.StringIO()
    csv.writer(out).writerows(filas)
    return send_file(
        io.BytesIO(out.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name="plantilla_trabajadores.csv",
    )


@app.route("/trabajadores/<int:id>/habitacion-aleatoria", methods=["POST"])
@admin_required
def habitacion_aleatoria_trabajador(id):
    """Asigna una habitación disponible aleatoria a un trabajador específico."""
    t = db.get_trabajador(id)
    if not t:
        abort(404)
    hab = db.asignar_habitacion_aleatoria(id)
    if hab:
        _regenerar_qr(id)
        t_nuevo = db.get_trabajador(id)
        flash(f"✅ Habitación asignada: Módulo {hab['modulo']} – Piso {hab['piso']} – N° {hab['numero']}", "success")
        if t_nuevo.get("email"):
            notificar_cambio_habitacion(t_nuevo, "Sin asignar",
                                        f"Módulo {hab['modulo']} – Piso {hab['piso']} – N° {hab['numero']}")
    else:
        flash("❌ No hay habitaciones disponibles para asignar.", "danger")
    return redirect(url_for("trabajador_detalle", id=id))


@app.route("/trabajadores/asignar-masivo", methods=["POST"])
@admin_required
def asignar_habitaciones_masivo():
    """Asigna habitaciones aleatorias a todos los trabajadores sin habitación."""
    sin_hab = db.get_trabajadores_sin_habitacion()
    asignados, sin_cupo = 0, 0
    for t in sin_hab:
        hab = db.asignar_habitacion_aleatoria(t["id"])
        if hab:
            _regenerar_qr(t["id"])
            asignados += 1
        else:
            sin_cupo += 1
    if asignados > 0:
        flash(f"✅ {asignados} habitaciones asignadas aleatoriamente.", "success")
    if sin_cupo > 0:
        flash(f"⚠️ {sin_cupo} trabajadores quedaron sin habitación (sin cupos disponibles).", "danger")
    if asignados == 0 and sin_cupo == 0:
        flash("ℹ️ Todos los trabajadores ya tienen habitación asignada.", "info")
    return redirect(url_for("trabajadores_lista"))


# ═══════════════════════════════════════════════════════
# HABITACIONES
# ═══════════════════════════════════════════════════════
@app.route("/habitaciones")
@admin_required
def habitaciones_mapa():
    habitaciones = db.get_todas_habitaciones()
    modulos = sorted(set(h["modulo"] for h in habitaciones))
    mapa = {}
    for m in modulos:
        pisos = sorted(set(h["piso"] for h in habitaciones if h["modulo"] == m))
        mapa[m] = {}
        for p in pisos:
            mapa[m][p] = [h for h in habitaciones if h["modulo"] == m and h["piso"] == p]
    # Trabajadores sin habitación (para el dropdown de asignación del mapa)
    trabajadores_disponibles = db.get_trabajadores_sin_habitacion()
    censo_hoy = db.get_censo_hoy_por_habitacion()
    return render_template("habitaciones/mapa.html", mapa=mapa, modulos=modulos,
                           trabajadores_disponibles=trabajadores_disponibles,
                           censo_hoy=censo_hoy)


@app.route("/habitaciones/asignar", methods=["POST"])
@admin_required
def asignar_habitacion():
    t_id_str = request.form.get("trabajador_id", "")
    if not t_id_str:
        flash("⚠️ Debes seleccionar un trabajador para asignar.", "warning")
        return redirect(url_for("habitaciones_mapa"))

    trabajador_id = int(t_id_str)
    hab_id = int(request.form["hab_id"])
    forzar = request.form.get("forzar") == "1"

    t = db.get_trabajador(trabajador_id)
    hab_anterior = f"Módulo {t.get('modulo')} – Piso {t.get('piso')} – Pieza {t.get('pieza')}" if t and t.get("modulo") else None

    # ── Verificar compatibilidad de turnos antes de asignar ─────────────────
    if not forzar and t.get("turno") in GRUPOS_TURNOS:
        grupo_nuevo = GRUPOS_TURNOS[t["turno"]]
        ocupantes_hab = db.get_ocupantes_habitacion(hab_id)
        turnos_incompatibles = []
        for ocu in ocupantes_hab:
            if ocu["id"] == trabajador_id:
                continue
            grupo_ocu = GRUPOS_TURNOS.get(ocu.get("turno", ""))
            if grupo_ocu and grupo_ocu != grupo_nuevo:
                turnos_incompatibles.append(f"{ocu['nombre']} ({ocu['turno']})")
        if turnos_incompatibles:
            nombres = ", ".join(turnos_incompatibles)
            flash(
                f"⚠️ Advertencia de cruce de turnos: la pieza ya tiene trabajadores de un grupo "
                f"incompatible ({nombres}). "
                f"Los turnos 14x14-A/D y 14x14-B/C no pueden mezclarse porque se verían "
                f"durante 7 días simultáneamente. "
                f"Confirma solo si es intencional.",
                "warning"
            )
            # Redirigir al detalle con parámetros para mostrar el formulario de confirmación
            return redirect(url_for(
                "trabajador_detalle",
                id=trabajador_id,
                confirmar_hab=hab_id,
                aviso="cruce_turno"
            ))

    ok = db.asignar_habitacion(hab_id, trabajador_id)
    if not ok:
        flash("❌ Esa habitación ya está completa para trabajadores activos. Elige otra.", "danger")
        return redirect(url_for("trabajador_detalle", id=trabajador_id))

    _regenerar_qr(trabajador_id)
    t_actualizado = db.get_trabajador(trabajador_id)
    hab_nueva = f"Módulo {t_actualizado.get('modulo')} – Piso {t_actualizado.get('piso')} – Pieza {t_actualizado.get('pieza')}"

    if hab_anterior and hab_anterior != hab_nueva and t_actualizado.get("email"):
        notificar_cambio_habitacion(t_actualizado, hab_anterior, hab_nueva)

    flash("✅ Habitación asignada exitosamente.", "success")
    return redirect(url_for("trabajador_detalle", id=trabajador_id))


@app.route("/habitaciones/nueva", methods=["POST"])
@admin_required
def nueva_habitacion():
    modulo = request.form["modulo"].strip().upper()
    try:
        piso = int(request.form.get("piso", 1))
        capacidad = int(request.form.get("capacidad", 3))
    except ValueError:
        flash("⚠️ Los valores de piso y capacidad deben ser números.", "warning")
        return redirect(url_for("habitaciones_mapa"))
        
    numero = request.form["numero"].strip()
    db.crear_habitacion(modulo, piso, numero, capacidad)
    flash(f"Habitación Módulo {modulo} – Piso {piso} – N° {numero} creada.", "success")
    return redirect(url_for("habitaciones_mapa"))


@app.route("/api/habitaciones/<int:hab_id>/ocupantes")
@admin_required
def api_ocupantes_habitacion(hab_id):
    """Retorna JSON con los ocupantes actuales de una habitación."""
    hab = db.get_habitacion(hab_id)
    if not hab:
        return jsonify({"error": "Not found"}), 404
    ocupantes = db.get_ocupantes_habitacion(hab_id)

    # Enriquecer cada ocupante con su estado de turno calculado hoy
    for t in ocupantes:
        if t.get("turno") and t.get("turno") in TURNOS:
            info = calcular_estado_turno(t["turno"], t.get("fecha_inicio_ciclo"))
            t["estado_turno_hoy"] = info["estado_calculado"]
            t["fecha_bajada"]     = info.get("fecha_bajada")
            t["fecha_retorno"]    = info.get("fecha_retorno")
            t["dias_restantes"]   = info.get("dias_restantes")
        else:
            t["estado_turno_hoy"] = None
            t["fecha_bajada"]     = None
            t["fecha_retorno"]    = None
            t["dias_restantes"]   = None

    return jsonify({
        "id": hab["id"],
        "modulo": hab["modulo"],
        "piso": hab["piso"],
        "numero": hab["numero"],
        "capacidad": hab.get("capacidad", 3),
        "ocupantes": ocupantes,
    })


@app.route("/habitaciones/<int:id>/estado", methods=["POST"])
@admin_required
def cambiar_estado_habitacion(id):
    nuevo = request.form["estado"]
    if nuevo == "Mantenimiento":
        # Liberar trabajador si hay uno
        hab = db.get_habitacion(id)
        if hab and hab.get("trabajador_id"):
            db.liberar_habitacion_de_trabajador(hab["trabajador_id"])
    db.actualizar_estado_habitacion(id, nuevo)
    flash("Estado de habitación actualizado.", "success")
    return redirect(url_for("habitaciones_mapa"))


@app.route("/habitaciones/<int:id>/eliminar", methods=["POST"])
@admin_required
def eliminar_habitacion(id):
    n = db.eliminar_habitacion(id)
    if n:
        flash(f"Pieza eliminada. {n} trabajador(es) desvinculado(s) de esa pieza.", "success")
    else:
        flash("Pieza eliminada.", "success")
    return redirect(url_for("habitaciones_mapa"))


@app.route("/habitaciones/modulo/renombrar", methods=["POST"])
@admin_required
def renombrar_modulo():
    actual = request.form.get("modulo_actual", "").strip().upper()
    nuevo  = request.form.get("modulo_nuevo",  "").strip().upper()
    if not actual or not nuevo:
        flash("Debes ingresar el nombre actual y el nuevo.", "warning")
        return redirect(url_for("habitaciones_mapa"))
    if actual == nuevo:
        flash("El nombre nuevo es igual al actual.", "warning")
        return redirect(url_for("habitaciones_mapa"))
    db.renombrar_modulo(actual, nuevo)
    flash(f"Módulo '{actual}' renombrado a '{nuevo}' correctamente.", "success")
    return redirect(url_for("habitaciones_mapa"))


@app.route("/habitaciones/modulo/eliminar", methods=["POST"])
@admin_required
def eliminar_modulo():
    modulo = request.form.get("modulo", "").strip().upper()
    if not modulo:
        flash("Módulo no especificado.", "warning")
        return redirect(url_for("habitaciones_mapa"))
    n = db.eliminar_modulo(modulo)
    flash(f"Módulo '{modulo}' eliminado — {n} pieza(s) removida(s). Trabajadores desvinculados de sus piezas.", "success")
    return redirect(url_for("habitaciones_mapa"))


@app.route("/habitaciones/modulo/agregar-piezas", methods=["POST"])
@admin_required
def agregar_piezas_modulo():
    modulo    = request.form.get("modulo", "").strip().upper()
    piso      = int(request.form.get("piso", 1))
    cantidad  = max(1, min(50, int(request.form.get("cantidad", 1))))
    capacidad = max(1, min(10, int(request.form.get("capacidad", 3))))
    n = db.agregar_piezas_modulo(modulo, piso, cantidad, capacidad)
    flash(f"✅ {n} pieza(s) agregadas al Módulo {modulo} · Piso {piso}.", "success")
    return redirect(url_for("habitaciones_mapa"))


# ═══════════════════════════════════════════════════════
# NOVEDADES
# ═══════════════════════════════════════════════════════
@app.route("/novedades")
@admin_required
def novedades_panel():
    novedades = db.get_novedades(limit=100)
    trabajadores = db.get_todos_trabajadores()
    return render_template("novedades/panel.html",
                           novedades=novedades, trabajadores=trabajadores)

# ═══════════════════════════════════════════════════════
# REPORTES Y EXCEL
# ═══════════════════════════════════════════════════════
@app.route("/reportes/asistencia")
@login_required
def reportes_asistencia():
    from datetime import date
    return render_template("reportes/asistencia.html", year=date.today().year, month=date.today().month)

@app.route("/api/reportes/asistencia/excel")
@login_required
def api_reportes_asistencia_excel():
    from datetime import date
    mes = int(request.args.get("mes", date.today().month))
    anio = int(request.args.get("anio", date.today().year))
    filtro_dia = request.args.get("dia", "")
    filtro_estado = request.args.get("estado", "")
    
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from calendar import monthrange
    from turnos import calcular_estado_turno
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Asistencia {mes:02d}-{anio}"
    
    fill_faena = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    fill_descanso = PatternFill(start_color="94A3B8", end_color="94A3B8", fill_type="solid")
    fill_vacaciones = PatternFill(start_color="A855F7", end_color="A855F7", fill_type="solid")
    fill_licencia = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
    fill_falla = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    
    font_bold_white = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    
    headers = ["Nombre", "RUT", "Cargo", "Turno"]
    _, num_days = monthrange(anio, mes)
    
    dias_a_mostrar = [int(filtro_dia)] if filtro_dia else list(range(1, num_days + 1))
    
    for d in dias_a_mostrar:
        headers.append(f"Día {d}" if filtro_dia else str(d))
    
    if not filtro_dia:
        headers.append("Total en Faena")
    
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_bold
        cell.alignment = center_align
    
    trabajadores = db.get_todos_trabajadores()
    novedades_mes = db.get_novedades_mes(anio, mes)
    
    nov_map = {}
    for n in novedades_mes:
        tid = n["trabajador_id"]
        d0 = date.fromisoformat(n["fecha_inicio"])
        d1 = date.fromisoformat(n["fecha_fin"]) if n["fecha_fin"] else date(2099, 12, 31)
        if tid not in nov_map:
            nov_map[tid] = []
        nov_map[tid].append((d0, d1, n["tipo"]))
    
    for t in trabajadores:
        if t["estado"] == "Desvinculado" and not t.get("fecha_inicio_ciclo"):
            continue
            
        fila = [t["nombre"], t["rut"], t.get("cargo", ""), t.get("turno", "")]
        dias_en_faena = 0
        estados_generados = []
        
        # Calcular los dias seleccionados
        for d in dias_a_mostrar:
            dia_actual = date(anio, mes, d)
            estado_dia = None
            if t["id"] in nov_map:
                for d0, d1, tipo in nov_map[t["id"]]:
                    if d0 <= dia_actual <= d1:
                        estado_dia = tipo
                        break
            
            if not estado_dia:
                if t.get("turno") and t.get("fecha_inicio_ciclo"):
                    info = calcular_estado_turno(t["turno"], t["fecha_inicio_ciclo"], target_date=dia_actual)
                    estado_dia = info["estado_calculado"]
                else:
                    estado_dia = "Indefinido"
                    
            estados_generados.append(estado_dia)
            if estado_dia == "Activo en campamento":
                dias_en_faena += 1
        
        # Filtrado de filas
        if filtro_estado:
            if filtro_estado == "Excepciones":
                valido = any(e in ["Falla", "Permiso"] for e in estados_generados)
            else:
                valido = any(e == filtro_estado for e in estados_generados)
            if not valido:
                continue

        fila.extend(estados_generados)
        if not filtro_dia:
            fila.append(dias_en_faena)
            
        ws.append(fila)
        
        current_row = ws.max_row
        for col_idx, est in enumerate(estados_generados, start=5):
            cell = ws.cell(row=current_row, column=col_idx)
            if est == "Activo en campamento":
                cell.value = "F"
                cell.fill = fill_faena
                cell.font = font_bold_white
            elif est == "En descanso":
                cell.value = "D"
                cell.fill = fill_descanso
                cell.font = font_bold_white
            elif est == "Vacaciones":
                cell.value = "V"
                cell.fill = fill_vacaciones
                cell.font = font_bold_white
            elif est == "Licencia Médica":
                cell.value = "L"
                cell.fill = fill_licencia
                cell.font = font_bold_white
            elif est in ["Falla", "Permiso", "Desvinculación"]:
                cell.value = "E"
                cell.fill = fill_falla
                cell.font = font_bold_white
            else:
                cell.value = ""
            cell.alignment = center_align
    
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"asistencia_{mes:02d}_{anio}.xlsx"
    )


# ═══════════════════════════════════════════════════════
# QR SCANNER (PORTERÍA)
# ═══════════════════════════════════════════════════════
@app.route("/porteria")
@admin_required
def porteria():
    return render_template("qr/scan.html")


@app.route("/api/qr/validar", methods=["POST"])
@admin_required
def api_validar_qr():
    data = request.get_json(force=True, silent=True) or {}
    token = (data.get("token") or "").strip()

    if not token:
        return jsonify({"ok": False, "error": "Token vacío"}), 400

    # 1. Verificar token en BD
    t = db.get_trabajador_by_token(token)
    if not t:
        return jsonify({
            "ok": False,
            "error": "QR inválido o revocado",
            "color": "red",
        }), 403

    # 2. Determinar tipo de movimiento (toggle Entrada/Salida)
    ultimo = db.get_ultimo_movimiento(t["id"])
    tipo = "Salida" if (ultimo and ultimo["tipo"] == "Entrada") else "Entrada"

    # 4. Registrar en historial
    db.registrar_movimiento(t["id"], tipo, "QR")

    # 5. Actualizar estado del trabajador
    nuevo_estado = "Activo en campamento" if tipo == "Entrada" else "En descanso"
    db.actualizar_estado_trabajador(t["id"], nuevo_estado)

    return jsonify({
        "ok": True,
        "tipo": tipo,
        "color": "green" if tipo == "Entrada" else "blue",
        "trabajador": {
            "nombre": t["nombre"],
            "rut": t["rut"],
            "cargo": t.get("cargo"),
            "turno": t.get("turno"),
            "modulo": t.get("modulo"),
            "piso": t.get("piso"),
            "pieza": t.get("pieza"),
        },
        "mensaje": f"✅ {tipo} registrada para {t['nombre']}",
    })


@app.route("/api/qr/censo", methods=["POST"])
@admin_required
def api_censo_qr():
    """Registra presencia de un trabajador durante censo nocturno/diurno."""
    data = request.get_json(force=True, silent=True) or {}
    token = (data.get("token") or "").strip()

    if not token:
        return jsonify({"ok": False, "error": "Token vacío"}), 400

    t = db.get_trabajador_by_token(token)
    if not t:
        return jsonify({"ok": False, "error": "QR inválido o revocado"}), 403

    hab_id = t.get("habitacion_id")
    usuario = session.get("username", "admin")
    db.registrar_censo(t["id"], hab_id, usuario)

    tiene_pieza = bool(t.get("modulo"))
    return jsonify({
        "ok": True,
        "tipo": "Censo",
        "trabajador": {
            "nombre": t["nombre"],
            "rut":    t["rut"],
            "cargo":  t.get("cargo"),
            "turno":  t.get("turno"),
            "modulo": t.get("modulo"),
            "piso":   t.get("piso"),
            "pieza":  t.get("pieza"),
        },
        "tiene_pieza": tiene_pieza,
    })


# ═══════════════════════════════════════════════════════
# REPORTES
# ═══════════════════════════════════════════════════════
@app.route("/reportes/movimientos")
@login_required
def reporte_movimientos():
    trabajador_id = request.args.get("trabajador_id", type=int)
    movimientos = db.get_movimientos(limit=500, trabajador_id=trabajador_id)
    trabajadores = db.get_todos_trabajadores()
    return render_template("reportes/movimientos.html",
                           movimientos=movimientos,
                           trabajadores=trabajadores,
                           trabajador_sel=trabajador_id)


# ═══════════════════════════════════════════════════════
# UTILIDADES INTERNAS
# ═══════════════════════════════════════════════════════
def _regenerar_qr(trabajador_id: int):
    """Regenera el token QR de un trabajador y lo guarda en BD."""
    t = db.get_trabajador(trabajador_id)
    if not t:
        return
    token = generar_token_qr(t)
    db.guardar_qr_token(trabajador_id, token)


# ═══════════════════════════════════════════════════════
# ADMINISTRACIÓN / DESARROLLO
# ═══════════════════════════════════════════════════════
@app.route("/admin/usuarios", methods=["GET", "POST"])
@admin_required
def admin_usuarios():
    from auth import hash_password
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        rol = request.form.get("rol", "viewer")
        nombre = request.form.get("nombre", "").strip()
        
        if not username or not password:
            flash("Usuario y contraseña son requeridos.", "warning")
            return redirect(url_for("admin_usuarios"))
            
        if db.usuario_existe(username):
            flash("Ese nombre de usuario ya existe.", "danger")
            return redirect(url_for("admin_usuarios"))
            
        db.crear_usuario(username, hash_password(password), rol, nombre)
        flash(f"Usuario {username} creado exitosamente.", "success")
        return redirect(url_for("admin_usuarios"))
        
    usuarios = db.get_todos_usuarios()
    return render_template("admin/usuarios.html", usuarios=usuarios)

@app.route("/admin/usuarios/<int:id>/eliminar", methods=["POST"])
@admin_required
def eliminar_usuario(id):
    u = db.get_usuario_by_id(id)
    if not u:
        abort(404)
    # Evitar que se elimine a si mismo
    if u['username'] == session.get('user'):
        flash("No puedes eliminar tu propio usuario mientras estás conectado.", "danger")
        return redirect(url_for("admin_usuarios"))
        
    db.eliminar_usuario(id)
    flash(f"Usuario {u['username']} eliminado.", "success")
    return redirect(url_for("admin_usuarios"))

@app.route("/admin/limpiar-bd", methods=["GET", "POST"])
@admin_required
def admin_limpiar_bd():
    if request.method == "POST":
        confirmacion = request.form.get("confirmacion", "").strip()
        if confirmacion == "LIMPIAR":
            db.limpiar_bd_desarrollo()
            flash("✅ Base de datos limpiada. Todos los trabajadores y habitaciones fueron eliminados.", "success")
            return redirect(url_for("dashboard"))
        flash("❌ Confirmación incorrecta. Escribe exactamente LIMPIAR.", "danger")
    return render_template("admin/limpiar_bd.html")


# ═══════════════════════════════════════════════════════
# ERROR HANDLERS
# ═══════════════════════════════════════════════════════
@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


if __name__ == "__main__":
    from init_data import init
    init()
    from database import migrar_db
    migrar_db()
    from config import PORT
    app.run(debug=True, host="0.0.0.0", port=PORT)
